""" Assumptions """
""" No new/depricated tests between different versions """
""" Will be fixed in later versions"""


from openpyxl import load_workbook	# for writing the report
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
import csv 							# for reading from csv log files
import os, fnmatch					# for finding the filename of the Report file
from collections import Counter		# for counting number of times a test is run
from statistics import mean

##################### variables to tune based on future changes #####################
apr_versions = ["VMS 4.0-8.05.05", "VMS-APR3.02-6.33.5", "VMS-3.1 - 6.55"]
latest_version = apr_versions[0]

### specify slots for different clients here
IPC_slots = [5]
VMS_slots = [6]
MKB_slots = [7] # mockingbird

slot_numbers = IPC_slots + VMS_slots + MKB_slots
##################### variables to tune based on future changes #####################

###################################### styles #######################################
light_green_fill	= PatternFill(fill_type='solid', start_color='FFDCEDC8', end_color='FFDCEDC8')
light_blue_fill		= PatternFill(fill_type='solid', start_color='FF4FC3F7', end_color='FF4FC3F7')
light_red_fill		= PatternFill(fill_type='solid', start_color='FFFFEBEE', end_color='FFFFEBEE')
pale_blue_fill  	= PatternFill(fill_type='solid', start_color='FFBBDEFB', end_color='FFBBDEFB')
green_fill 			= PatternFill(fill_type='solid', start_color='FF689F38', end_color='FF689F38')
red_fill 			= PatternFill(fill_type='solid', start_color='FFF44336', end_color='FFF44336')
blue_fill 			= PatternFill(fill_type='solid', start_color='FF1976D2', end_color='FF1976D2')
center_alignment 	= Alignment(horizontal="center", vertical="center")
###################################### styles #######################################

############################### data sheet functions ################################
from itertools import count, product, islice
from string import ascii_uppercase
def multiletters(seq, start):
	found = False
	for n in count(1):
		for s in product(seq, repeat=n):
			result = ''.join(s)
			if result == start:
				found = True
			if not found:
				continue
			yield result

def char_range(c1, length):
    return list(islice(multiletters(ascii_uppercase, c1), length))

def read_csv(apr_version, slot_number, log_file_directories):
	try:
		log_file_contents = []
		log_file_name = '{}/{}/logs{}.csv'.format(log_file_directories, str(apr_version), str(slot_number))
		with open(log_file_name, newline='') as csvfile:
			spamreader = csv.reader(csvfile, delimiter=',', quotechar='|')
			for row in spamreader:
				if row[3] == "PASS":
					row[2] = float(row[2])
					log_file_contents.append(row)
		return log_file_contents
	except OSError as err:
		print("\n\nThe file {} does not exists!!!\n\n".format(log_file_name))
		raise

def collect_from_csv(log_file_directories):
	# read and store date from log files
	IPC_records = {apr_version: [] for apr_version in apr_versions}
	VMS_records = {apr_version: [] for apr_version in apr_versions}
	MKB_records = {apr_version: [] for apr_version in apr_versions}
	for apr_version in apr_versions:
		for slot_number in slot_numbers:
			log_file_contents = read_csv(apr_version, slot_number, log_file_directories)
			if slot_number in IPC_slots:
				IPC_records[apr_version].extend(log_file_contents)
			elif slot_number in VMS_slots:
				VMS_records[apr_version].extend(log_file_contents)
			else: # MKB
				MKB_records[apr_version].extend(log_file_contents)
	return IPC_records, VMS_records, MKB_records

def pad(records, tc, tc_name, num_to_pad):
	selected_records = [record for record in records if record[0] == tc]
	avg = mean([float(selected_record[2]) for selected_record in selected_records])
	pad_record = [tc, tc_name, str(avg), "PADDED_DATA", "PADDED_DATA"]
	return records + num_to_pad * [pad_record]


def check_and_pad(IPC_records, VMS_records, MKB_records):
	for apr_version in apr_versions:
		r1 = IPC_records[apr_version]
		r2 = VMS_records[apr_version]
		r3 = MKB_records[apr_version]
		counter1 = Counter([(elem[0], elem[1]) for elem in r1])
		counter2 = Counter([(elem[0], elem[1]) for elem in r2])
		counter3 = Counter([(elem[0], elem[1]) for elem in r3])
		for tc, tc_name in counter1:
			if counter1[(tc, tc_name)] > counter2[(tc, tc_name)] and counter1[(tc, tc_name)] > counter3[(tc, tc_name)]:
				VMS_records[apr_version] = pad(VMS_records[apr_version], tc, tc_name, counter1[(tc, tc_name)] - counter2[(tc, tc_name)])
				MKB_records[apr_version] = pad(MKB_records[apr_version], tc, tc_name, counter1[(tc, tc_name)] - counter3[(tc, tc_name)])
			elif counter2[(tc, tc_name)] > counter1[(tc, tc_name)] and counter2[(tc, tc_name)] > counter3[(tc, tc_name)]:
				IPC_records[apr_version] = pad(IPC_records[apr_version], tc, tc_name, counter2[(tc, tc_name)] - counter1[(tc, tc_name)])
				MKB_records[apr_version] = pad(MKB_records[apr_version], tc, tc_name, counter2[(tc, tc_name)] - counter3[(tc, tc_name)])
			elif counter3[(tc, tc_name)] > counter1[(tc, tc_name)] and counter3[(tc, tc_name)] > counter2[(tc, tc_name)]:
				IPC_records[apr_version] = pad(IPC_records[apr_version], tc, tc_name, counter3[(tc, tc_name)] - counter1[(tc, tc_name)])
				VMS_records[apr_version] = pad(VMS_records[apr_version], tc, tc_name, counter3[(tc, tc_name)] - counter2[(tc, tc_name)])

	return IPC_records, VMS_records, MKB_records

def write_header(ws, ord_mac):
	for ver_count, apr_version in enumerate(apr_versions):
		ws["{}1".format(chr(ord_mac + ver_count))] = apr_versions[ver_count]
		if ver_count != 0:
			ws["{}1".format(chr(ord_mac + len(apr_versions) + 2 * ver_count - 2))] = \
				"Difference\n({}-{})".format(apr_versions[ver_count], latest_version)
			ws["{}1".format(chr(ord_mac + len(apr_versions) + 2 * ver_count - 1))] = \
				"%Change"

def write_test_data(ws, ord_mac, records, num_records):
	for i in range(num_records):
		for ver_count, apr_version in enumerate(apr_versions):
			ws["{}{}".format(chr(ord_mac + ver_count), i+2)] = records[apr_version][i][2]
			if ver_count != 0:
				ws["{}{}".format(chr(ord_mac + len(apr_versions) + 2 * ver_count - 2), i+2)] = \
					"={0}{2} - {1}{2}".format(chr(ord_mac + ver_count), chr(ord_mac), i + 2)
				ws["{}{}".format(chr(ord_mac + len(apr_versions) + 2 * ver_count - 1), i+2)] = \
					"={0}{2}/{1}{2}*100".format(chr(ord_mac + len(apr_versions) + 2 * ver_count - 2), chr(ord_mac), i + 2)

def write_contents(ws, IPC_records, VMS_records, MKB_records):
	### Find the ASCII number corresponding to the starting column letter
	red_zone_width = len(apr_versions) + 2 * (len(apr_versions)-1)
	ord_VMS = ord("C")
	ord_IPC = ord_VMS + red_zone_width
	ord_MKB = ord_IPC + red_zone_width

	### write the headers for each column
	ws['A1'], ws['B1'] = "TC", "TC Name"
	write_header(ws, ord_VMS)
	write_header(ws, ord_IPC)
	write_header(ws, ord_MKB)

	### Write the contents of the first 2 columns
	first_key = list(VMS_records.keys())[0]
	num_records = len(VMS_records[first_key])
	for i in range(num_records):
		ws["{}{}".format("A", i+2)] = VMS_records[first_key][i][0]
		ws["{}{}".format("B", i+2)] = VMS_records[first_key][i][1]

	### Write the test data
	write_test_data(ws, ord_VMS, VMS_records, num_records)
	write_test_data(ws, ord_IPC, IPC_records, num_records)
	write_test_data(ws, ord_MKB, MKB_records, num_records)
	
def paint_and_format_data(ws, ord_mac, red_zone_width, paint, i):
	for letter in char_range(chr(ord_mac), red_zone_width):
		ws['{}{}'.format(letter, i)].fill = paint
		ws['{}{}'.format(letter, i)].alignment = center_alignment
		ws['{}{}'.format(letter, i)].number_format = '0.00'

def format_sheet(ws):
	red_zone_width = len(apr_versions) + 2 * (len(apr_versions)-1)
	ord_VMS = ord("C")
	ord_IPC = ord_VMS + red_zone_width
	ord_MKB = ord_IPC + red_zone_width

	for letter in char_range("A", red_zone_width+2):
		ws['{}1'.format(letter)].fill = green_fill
	for letter in char_range(chr(ord_IPC), red_zone_width):
		ws['{}1'.format(letter)].fill = red_fill
	for letter in char_range(chr(ord_MKB), red_zone_width):
		ws['{}1'.format(letter)].fill = blue_fill
	for letter in char_range("A", red_zone_width * 3 + 2):
		ws.column_dimensions[letter].width = 15

	for i in range(2, ws.max_row+1):
		paint_and_format_data(ws, ord_VMS, red_zone_width, light_green_fill, i)
		paint_and_format_data(ws, ord_IPC, red_zone_width, light_red_fill, i)
		paint_and_format_data(ws, ord_MKB, red_zone_width, pale_blue_fill, i)
		if i == 2 or ws["A{}".format(i)].value != ws["A{}".format(i-1)].value:
			ws['A{}'.format(i)].fill = light_blue_fill
			ws['B{}'.format(i)].fill = light_blue_fill
	
	ws.row_dimensions[1].height = 60
	ws.column_dimensions["A"].width = 6
	ws.column_dimensions["B"].width = 50

def write_report_data_page(ws, IPC_records, VMS_records, MKB_records):
	write_contents(ws, IPC_records, VMS_records, MKB_records)
	format_sheet(ws)
############################### data sheet functions ################################

############################### front page functions ################################
def write_header_FP(ws):
	ord_latest = ord("C")
	ws.merge_cells("A1:A4")
	ws.merge_cells("B1:B4")

	chars = char_range("C", 6*len(apr_versions)+1)
	for i in range(len(apr_versions)):
		ws.merge_cells('{}1:{}1'.format(chars[0+6*i], chars[5+6*i]))
		ws.merge_cells('{}2:{}2'.format(chars[0+6*i], chars[5+6*i]))
		ws.merge_cells('{}3:{}3'.format(chars[0+6*i], chars[1+6*i]))
		ws.merge_cells('{}3:{}3'.format(chars[2+6*i], chars[3+6*i]))
		ws.merge_cells('{}3:{}3'.format(chars[4+6*i], chars[5+6*i]))

	ws.merge_cells('{0}1:{0}4'.format(chars[6*len(apr_versions)]))

	n = len(apr_versions)
	apr_versions_to_write = [[apr_version] + 5 * [None] for apr_version in apr_versions]
	apr_versions_to_write = [item for sublist in apr_versions_to_write for item in sublist]
	rows = [
    	['Test Case ID', 'Test Case Name'] + apr_versions_to_write + ["Steps"],
    	[None, None] + ["UP Time: 1 Hour-DVR: 1% full", None, None, None, None, None] * n,
    	[None, None] + ["VMS", None, "IPC", None, "Mockingbird", None] * n,
    	[None, None] + ["Mean(Sec)","Standard Deviation( σ)"] * 3 * n,
	]
	for row in rows:
	    ws.append(row)

def read_from_des_txt(tc):
	f = open('Input/{}.txt'.format(tc), 'r')
	result = f.read()
	f.close()
	return result

def write_front_page(front_ws, data_ws):
	write_header_FP(front_ws)

	test_head_row_numbers = []
	tc_s = []
	tc_names = []
	for i in range(2, data_ws.max_row+1):
		if i == 2 or i == data_ws.max_row\
		or data_ws["A{}".format(i)].value != data_ws["A{}".format(i-1)].value:
			tc_s.append(data_ws["A{}".format(i)].value)
			tc_names.append(data_ws["B{}".format(i)].value)
			test_head_row_numbers.append(i)
	
	num_vers = len(apr_versions)
	chars = char_range("C", 3 * 3 * num_vers)
	red_zone_width = num_vers + 2 * (num_vers-1)
	VMS_chars = chars[:num_vers]
	IPC_chars = chars[red_zone_width : red_zone_width + num_vers]
	MKB_chars = chars[2 * red_zone_width : 2 * red_zone_width + num_vers]

	for i in range(len(tc_s)-1):
		row = [tc_s[i], tc_names[i]]
		for j in range(num_vers):
			row.append("=AVERAGE(PerformanceDATA!{0}{1}:{0}{2})".\
				format(VMS_chars[j], test_head_row_numbers[i], test_head_row_numbers[i+1]))
			row.append("=STDEV(PerformanceDATA!{0}{1}:{0}{2})".\
				format(VMS_chars[j], test_head_row_numbers[i], test_head_row_numbers[i+1]))
			row.append("=AVERAGE(PerformanceDATA!{0}{1}:{0}{2})".\
				format(IPC_chars[j], test_head_row_numbers[i], test_head_row_numbers[i+1]))
			row.append("=STDEV(PerformanceDATA!{0}{1}:{0}{2})".\
				format(IPC_chars[j], test_head_row_numbers[i], test_head_row_numbers[i+1]))
			row.append("=AVERAGE(PerformanceDATA!{0}{1}:{0}{2})".\
				format(MKB_chars[j], test_head_row_numbers[i], test_head_row_numbers[i+1]))
			row.append("=STDEV(PerformanceDATA!{0}{1}:{0}{2})".\
				format(MKB_chars[j], test_head_row_numbers[i], test_head_row_numbers[i+1]))
		row.append(read_from_des_txt(tc_s[i]))
		front_ws.append(row)
############################### front page functions ################################

################################### main function ###################################
def write_report():
	# extract records from log files
	IPC_records, VMS_records, MKB_records = collect_from_csv("logs")
	# sort each list in each dict by the first column
	for apr_version in apr_versions:
		IPC_records[apr_version] = sorted(IPC_records[apr_version], key=lambda elem: elem[0])
		VMS_records[apr_version] = sorted(VMS_records[apr_version], key=lambda elem: elem[0])
		MKB_records[apr_version] = sorted(MKB_records[apr_version], key=lambda elem: elem[0])
	# Check data and pad if necessary
	IPC_records, VMS_records, MKB_records = check_and_pad(IPC_records, VMS_records, MKB_records)

	# write to the report file
	wb = Workbook()
	front_ws = wb.active
	front_ws.title = "Dashboard"
	data_ws = wb.create_sheet(title = 'PerformanceDATA')

	################################################################################
	write_report_data_page(data_ws, IPC_records, VMS_records, MKB_records)
	write_front_page(front_ws, data_ws)
	################################################################################
	wb.save('Report Output/output_report.xlsx')

################################### main function ###################################
if __name__ == '__main__':
    write_report()