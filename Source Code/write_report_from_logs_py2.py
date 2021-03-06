from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.cell import get_column_letter
from openpyxl.chart import Reference, Series, BarChart3D
from openpyxl.chart.layout import Layout, ManualLayout
import sys
import csv
import os, fnmatch
from collections import Counter
from itertools import chain, count, product, islice
from datetime import datetime, date, time
from dateutil import parser
from string import ascii_uppercase

##################### variables to tune based on future changes #####################
def read_metadata(filename):
	with open(filename, 'rb') as csvfile:
		spamreader = csv.reader(csvfile, delimiter=',', quotechar='|')
		return next(spamreader)
def read_apr_versions_and_slots(filename):
	apr_versions = []
	slots_in_apr = {}
	with open(filename, 'rb') as csvfile:
		spamreader = csv.reader(csvfile, delimiter=',', quotechar='|')
		for row in spamreader:
			slots_in_apr[row[0]] = [int(s) for s in row[1:] if s != ""]
			apr_versions.append(row[0])
	return apr_versions, slots_in_apr
try:
	VMS_slots = read_metadata("Metadata/VMS_slots.csv")
	VMS_slots = [int(s) for s in VMS_slots]
except OSError as err:
	print("Caution!!!\nFile 'VMS_slots' not found!\nDefault values used!\nPlease see Readme.md for more information")
	VMS_slots = [1,3,6,9,11,13,15]
try:
	IPC_slots = read_metadata("Metadata/IPC_slots.csv")
	IPC_slots = [int(s) for s in IPC_slots]
except OSError as err:
	print("Caution!!!\nFile 'IPC_slots' not found!\nDefault values used!\nPlease see Readme.md for more information")
	IPC_slots = [2,4,10,12,14,16]
try:
	MKB_slots = read_metadata("Metadata/MKB_slots.csv")
	MKB_slots = [int(s) for s in MKB_slots]
except OSError as err:
	print("Caution!!!\nFile 'MKB_slots' not found!\nDefault values used!\nPlease see Readme.md for more information")
	MKB_slots = [6,7,8] # mockingbird
try:
	apr_versions, slots_in_apr = read_apr_versions_and_slots("Metadata/Version_names.csv")
except OSError as err:
	print("Caution!!!\nFile 'Version_names' not found!\nDefault values used!\nPlease see Readme.md for more information")
	apr_versions = ["VMS 4.0-8.05.05", "VMS-APR3.02-6.33.5", "VMS-3.1 - 6.55"]
	slots_in_apr = {
		"VMS 4.0-8.05.05": [4,5,6,7,8,9,10,11,12], 
		"VMS-APR3.02-6.33.5": [13,14,15,16], 
		"VMS-3.1 - 6.55": [1,2,3]
	}
try:
	cutoff_length = int(read_metadata("Metadata/cutoff_length.csv")[0])
except OSError as err:
	print("Caution!!!\nFile 'cutoff_length' not found!\nDefault values used!\nPlease see Readme.md for more information")
	cutoff_length = 25

##################### variables to tune based on future changes #####################

###################################### styles #######################################
pale_green_fill		= PatternFill(fill_type='solid', start_color='FFDCEDC8', end_color='FFDCEDC8')
pale_red_fill		= PatternFill(fill_type='solid', start_color='FFFFEBEE', end_color='FFFFEBEE')
pale_blue_fill  	= PatternFill(fill_type='solid', start_color='FFBBDEFB', end_color='FFBBDEFB')
pale_purple_fill	= PatternFill(fill_type='solid', start_color='FFD1C4E9', end_color='FFD1C4E9')
light_blue_fill		= PatternFill(fill_type='solid', start_color='FF4FC3F7', end_color='FF4FC3F7')
green_fill 			= PatternFill(fill_type='solid', start_color='FF81C784', end_color='FF81C784')
red_fill 			= PatternFill(fill_type='solid', start_color='FFE57373', end_color='FFE57373')
blue_fill 			= PatternFill(fill_type='solid', start_color='FF7986CB', end_color='FF7986CB')
orange_fill			= PatternFill(fill_type='solid', start_color='FFFB8C00', end_color='FFFB8C00')
grey_fill 			= PatternFill(fill_type='solid', start_color='FFECEFF1', end_color='FFECEFF1')
metalic_green_fill	= PatternFill(fill_type='solid', start_color='FF80CBC4', end_color='FF80CBC4')
center_alignment 	= Alignment(horizontal="center", vertical="center")
small_font			= Font(size=8)
black_side  		= Side(border_style='thin', color='FF000000')
black_border		= Border(left=black_side, right=black_side, top=black_side, bottom=black_side)
###################################### styles #######################################

################################## test functions ###################################
def see_counters(VMS_records, IPC_records, MKB_records, maximums):
	counters1, counters2, counters3 = [],[],[]
	for apr_version in apr_versions:
		r1 = VMS_records[apr_version]
		r2 = IPC_records[apr_version]
		r3 = MKB_records[apr_version]
		counters1.append(Counter([record[0] for record in r1]))
		counters2.append(Counter([record[0] for record in r2]))
		counters3.append(Counter([record[0] for record in r3]))
	for i in range(len(apr_versions)):
		print(counters1[i])
	print()
	for i in range(len(apr_versions)):
		print(counters2[i])
	print()
	for i in range(len(apr_versions)):
		print(counters3[i])
	print()
	print(maximums)
################################## test functions ###################################

############################### data sheet functions ################################
latest_version = apr_versions[0]
slot_numbers = IPC_slots + VMS_slots + MKB_slots

def mean(lst):
	return(sum(lst)/float(len(lst)))

def multiletters(seq, start):
	found = False
	for n in count(1):
		for s in product(seq, repeat=n):
			result = ''.join(s)
			if result == start:
				found = True
			if not found:
				continue
			yield result

def char_range(c1, length):
    return list(islice(multiletters(ascii_uppercase, c1), length))\

def column_num_to_letter(column_num):
	return list(char_range("A", column_num))[-1]

def read_csv(slot_number, log_file_directories):
	try:
		log_file_contents = []
		log_file_name = '{}/logs{}.csv'.format(log_file_directories, str(slot_number))
		with open(log_file_name, 'rb') as csvfile:
			spamreader = csv.reader(csvfile, delimiter=',', quotechar='|')
			for row in spamreader:
				if len(row) != 0 and len(row) > 4 and row[3] == "PASS":
					row[2] = float(row[2])
					row[4] = parser.parse(row[4])
					log_file_contents.append(row)
		return log_file_contents
	except OSError as err:
		print("\n\nThe file {} does not exists!!!\n\n".format(log_file_name))
		raise

def collect_from_csv(log_file_directories):
	# read and store date from log files
	IPC_records = {apr_version: [] for apr_version in apr_versions}
	VMS_records = {apr_version: [] for apr_version in apr_versions}
	MKB_records = {apr_version: [] for apr_version in apr_versions}
	for apr_version in apr_versions:
		for slot_number in slot_numbers:
			if slot_number not in slots_in_apr[apr_version]:
				continue
			log_file_contents = read_csv(slot_number, log_file_directories)
			if slot_number in IPC_slots:
				IPC_records[apr_version].extend(log_file_contents)
			elif slot_number in VMS_slots:
				VMS_records[apr_version].extend(log_file_contents)
			else: # MKB
				MKB_records[apr_version].extend(log_file_contents)
	return IPC_records, VMS_records, MKB_records

def pad(records, tc, tc_name, num_to_pad):
	# pad a specific amount of records
	# Rules: if there are existing records under same machine^apr_version^tc, take the average to pad
	# More Rule: if there no existing records, pad with value 0
	selected_records = [record for record in records if record[0] == tc]
	if len(selected_records) != 0:
		avg = mean([float(selected_record[2]) for selected_record in selected_records])
	else: 
		avg = 0.0
	pad_record = [tc, tc_name, float(avg), "PADDED_DATA", "PADDED_DATA"]
	return records + num_to_pad * [pad_record]

def check_and_pad(VMS_records, IPC_records, MKB_records):
	# Find all the test case names
	all_test_cases = set()
	for apr_version in apr_versions:
		for record in VMS_records[apr_version]:
			all_test_cases.update([(record[0], record[1])])
		for record in IPC_records[apr_version]:
			all_test_cases.update([(record[0], record[1])])
		for record in MKB_records[apr_version]:
			all_test_cases.update([(record[0], record[1])])
	all_test_cases = sorted(list(all_test_cases), key=lambda elem: elem[0])

	# Counters for later use
	VMS_counters = []
	IPC_counters = []
	MKB_counters = []
	for apr_version in apr_versions:
		r1 = VMS_records[apr_version]
		r2 = IPC_records[apr_version]
		r3 = MKB_records[apr_version]
		VMS_counters.append(Counter([(elem[0], elem[1]) for elem in r1]))
		IPC_counters.append(Counter([(elem[0], elem[1]) for elem in r2]))
		MKB_counters.append(Counter([(elem[0], elem[1]) for elem in r3]))

	# find the max number of records for that tc across machines and Versions for each test cases
	maximums = {(tc, tc_name): 0 for tc, tc_name in all_test_cases}
	for tc, tc_name in all_test_cases:
		for apr_version in apr_versions:
			for counter in VMS_counters:
				maximums[(tc, tc_name)] = max(maximums[(tc, tc_name)], counter[(tc, tc_name)])
			for counter in IPC_counters:
				maximums[(tc, tc_name)] = max(maximums[(tc, tc_name)], counter[(tc, tc_name)])
			for counter in MKB_counters:
				maximums[(tc, tc_name)] = max(maximums[(tc, tc_name)], counter[(tc, tc_name)])
	
	# Padding happens here
	# for each test case name
	for tc, tc_name in all_test_cases:
		# for each apr_version
		for i, apr_version in enumerate(apr_versions):
			# for each machine, pad the records to max number
			num_to_pad = maximums[(tc, tc_name)] - VMS_counters[i][(tc, tc_name)]
			VMS_records[apr_version] = pad(VMS_records[apr_version], tc, tc_name, num_to_pad)
			num_to_pad = maximums[(tc, tc_name)] - IPC_counters[i][(tc, tc_name)]
			IPC_records[apr_version] = pad(IPC_records[apr_version], tc, tc_name, num_to_pad)
			num_to_pad = maximums[(tc, tc_name)] - MKB_counters[i][(tc, tc_name)]
			MKB_records[apr_version] = pad(MKB_records[apr_version], tc, tc_name, num_to_pad)	
	return IPC_records, VMS_records, MKB_records, maximums

def trim_data(records, cutoff_length = 25):
	records_results = {apr_version: [] for apr_version in apr_versions}
	all_test_cases = set()
	for apr_version in apr_versions:
		for record in records[apr_version]:
			all_test_cases.update([record[0]])
	for apr_version in apr_versions:
		for tc in all_test_cases:
			selected_records = [record for record in records[apr_version] if record[0] == tc]
			selected_records = sorted(selected_records, key=lambda elem: elem[4], reverse = True)
			records_results[apr_version].extend	(selected_records[:cutoff_length])
	return records_results

def write_header(ws, column_num):
	# write the header of the data page
	for ver_count, apr_version in enumerate(apr_versions):
		ws.cell(row = 1, column = column_num + ver_count).value = apr_versions[ver_count]
		if ver_count != 0:
			ws.cell(row = 1, column = column_num + len(apr_versions) + 2 * ver_count - 2).value = \
				"Difference\n({}-{})".format(apr_versions[ver_count], latest_version)
			ws.cell(row = 1, column = column_num + len(apr_versions) + 2 * ver_count - 1).value = \
				"%Change"

def write_test_data(ws, column_num, records, num_records):
	# write, row by row, the test data onto the data page
	for i in range(num_records):
		for ver_count, apr_version in enumerate(apr_versions):
			ws.cell(row = i+2, column = column_num + ver_count).value = records[apr_version][i][2]
			if ver_count != 0:
				column1 = column_num_to_letter(column_num + ver_count)
				column2 = column_num_to_letter(column_num)
				column3 = column_num_to_letter(column_num + len(apr_versions) + 2 * ver_count - 2)
				ws.cell(row = i+2, column = column_num + len(apr_versions) + 2 * ver_count - 2).value = \
					"={0}{2} - {1}{2}".format(column1, column2, i + 2)
				ws.cell(row = i+2, column = column_num + len(apr_versions) + 2 * ver_count - 1).value = \
					"={0}{2}/{1}{2}*100".format(column3, column2, i + 2)

def write_contents(ws, IPC_records, VMS_records, MKB_records):
	### Find the ASCII number corresponding to the starting column letter
	red_zone_width = len(apr_versions) + 2 * (len(apr_versions)-1)
	column_VMS = 3
	column_IPC = column_VMS + red_zone_width
	column_MKB = column_IPC + red_zone_width

	### write the headers for each column
	ws['A1'], ws['B1'] = "TC", "TC Name"
	write_header(ws, column_VMS)
	write_header(ws, column_IPC)
	write_header(ws, column_MKB)

	### Write the contents of the first 2 columns
	first_key = list(VMS_records.keys())[0]
	num_records = len(VMS_records[first_key])
	for i in range(num_records):
		ws["{}{}".format("A", i+2)] = VMS_records[first_key][i][0]
		ws["{}{}".format("B", i+2)] = VMS_records[first_key][i][1]

	write_test_data(ws, column_VMS, VMS_records, num_records)
	write_test_data(ws, column_IPC, IPC_records, num_records)
	write_test_data(ws, column_MKB, MKB_records, num_records)

### style format functions for data page
def paint_and_format_data(ws, column_num, red_zone_width, paint, i):
	for letter in char_range(column_num_to_letter(column_num), red_zone_width):
		ws['{}{}'.format(letter, i)].fill = paint
		ws['{}{}'.format(letter, i)].alignment = center_alignment
		ws['{}{}'.format(letter, i)].number_format = '0.00'

def format_sheet(ws):
	red_zone_width = len(apr_versions) + 2 * (len(apr_versions)-1)
	column_VMS = 3
	column_IPC = column_VMS + red_zone_width
	column_MKB = column_IPC + red_zone_width

	# paint header
	for letter in char_range("A", red_zone_width+2):
		ws['{}1'.format(letter)].fill = green_fill
	for letter in char_range(column_num_to_letter(column_IPC), red_zone_width):
		ws['{}1'.format(letter)].fill = red_fill
	for letter in char_range(column_num_to_letter(column_MKB), red_zone_width):
		ws['{}1'.format(letter)].fill = blue_fill
	for letter in char_range("A", red_zone_width * 3 + 2):
		ws.column_dimensions[letter].width = 15

	# paint data
	for i in range(2, ws.max_row+1):
		paint_and_format_data(ws, column_VMS, red_zone_width, pale_green_fill, i)
		paint_and_format_data(ws, column_IPC, red_zone_width, pale_red_fill, i)
		paint_and_format_data(ws, column_MKB, red_zone_width, pale_blue_fill, i)
		if i == 2 or ws["A{}".format(i)].value != ws["A{}".format(i-1)].value:
			ws['A{}'.format(i)].fill = light_blue_fill
			ws['B{}'.format(i)].fill = light_blue_fill
	
	# set alignment and border
	for i in range(1, ws.max_column+1):
		letter = get_column_letter(i)
		for j in range(1, ws.max_row+1):
			ws["{}{}".format(letter,j)].border = black_border
			if j == 1:
				ws["{}{}".format(letter,j)].alignment = Alignment(wrapText=True)
	
	# set dimensions
	ws.row_dimensions[1].height = 60
	ws.column_dimensions["A"].width = 6
	ws.column_dimensions["B"].width = 50

def write_report_data_page(ws, IPC_records, VMS_records, MKB_records):
	write_contents(ws, IPC_records, VMS_records, MKB_records)
	format_sheet(ws)
############################### data sheet functions ################################

############################### front page functions ################################
def write_header_FP(ws):
	# Write the Front page header
	ws.merge_cells("A1:A4")
	ws.merge_cells("B1:B4")

	chars = char_range("C", 6*len(apr_versions)+1)
	for i in range(len(apr_versions)):
		ws.merge_cells('{}1:{}1'.format(chars[0+6*i], chars[5+6*i]))
		ws.merge_cells('{}2:{}2'.format(chars[0+6*i], chars[5+6*i]))
		ws.merge_cells('{}3:{}3'.format(chars[0+6*i], chars[1+6*i]))
		ws.merge_cells('{}3:{}3'.format(chars[2+6*i], chars[3+6*i]))
		ws.merge_cells('{}3:{}3'.format(chars[4+6*i], chars[5+6*i]))

	ws.merge_cells('{0}1:{0}4'.format(chars[6*len(apr_versions)]))

	n = len(apr_versions)
	apr_versions_to_write = [[apr_version] + 5 * [None] for apr_version in apr_versions]
	apr_versions_to_write = [item for sublist in apr_versions_to_write for item in sublist]
	rows = [
    	['Test \nCase ID', 'Test Case Name'] + apr_versions_to_write + ["Steps"],
    	[None, None] + ["UP Time: 1 Hour-DVR: 1% full", None, None, None, None, None] * n,
    	[None, None] + ["VMS", None, "IPC", None, "Mockingbird", None] * n,
    	[None, None] + ["Mean(Sec)","STDEV(Sec)"] * 3 * n,
	]
	for row in rows:
	    ws.append(row)

def read_from_des_txt(tc):
	# Open the corresponding file for the tc and read the test description
	try:
		f = open('Test Descriptions/{}.txt'.format(tc), 'r')
		result = f.read()
		f.close()
	except OSError:
		result = "No description specified"
	return result

def write_front_page_data(front_ws, data_ws, maximums):\
	# Write the data section in the front page
	test_head_row_numbers = []
	tc_s = []
	tc_names = []
	cumulative_row_number = 2
	for item in sorted(maximums):
		tc_s.append(item[0])
		tc_names.append(item[1])
		test_head_row_numbers.append(cumulative_row_number)
		cumulative_row_number += maximums[item]
	test_head_row_numbers.append(cumulative_row_number)

	num_vers = len(apr_versions)
	chars = char_range("C", 3 * 3 * num_vers)
	red_zone_width = num_vers + 2 * (num_vers-1)
	VMS_chars = chars[:num_vers]
	IPC_chars = chars[red_zone_width : red_zone_width + num_vers]
	MKB_chars = chars[2 * red_zone_width : 2 * red_zone_width + num_vers]

	for i in range(len(tc_s)-1):
		row = [tc_s[i], tc_names[i]]
		for j in range(num_vers):
			row.append("=AVERAGE(PerformanceDATA!{0}{1}:{0}{2})".\
				format(VMS_chars[j], test_head_row_numbers[i], test_head_row_numbers[i+1]))
			row.append("=STDEV(PerformanceDATA!{0}{1}:{0}{2})".\
				format(VMS_chars[j], test_head_row_numbers[i], test_head_row_numbers[i+1]))
			row.append("=AVERAGE(PerformanceDATA!{0}{1}:{0}{2})".\
				format(IPC_chars[j], test_head_row_numbers[i], test_head_row_numbers[i+1]))
			row.append("=STDEV(PerformanceDATA!{0}{1}:{0}{2})".\
				format(IPC_chars[j], test_head_row_numbers[i], test_head_row_numbers[i+1]))
			row.append("=AVERAGE(PerformanceDATA!{0}{1}:{0}{2})".\
				format(MKB_chars[j], test_head_row_numbers[i], test_head_row_numbers[i+1]))
			row.append("=STDEV(PerformanceDATA!{0}{1}:{0}{2})".\
				format(MKB_chars[j], test_head_row_numbers[i], test_head_row_numbers[i+1]))
		row.append(read_from_des_txt(tc_s[i]))
		front_ws.append(row)

def format_FP(front_ws):
	# Formatting of the front page
	# What a mess... Might need to refactor it in the future
	for i in range(1, front_ws.max_column+1):
		letter = get_column_letter(i)
		num_vers = len(apr_versions)
		if i in range(1, front_ws.max_column):
			for j in range(1, front_ws.max_row+1):
				front_ws["{}{}".format(letter, j)].alignment = center_alignment
		if i in [1, 2, 3 + 6 * num_vers]:
			for j in range(1, front_ws.max_row+1):
				front_ws["{}{}".format(letter, j)].fill = pale_green_fill
		if i in range(3, 3 + 6 * num_vers, 2):
			front_ws["{}1".format(letter)].fill = orange_fill
			front_ws["{}2".format(letter)].fill = grey_fill
			for j in range(5, front_ws.max_row+1):
				front_ws["{}{}".format(letter, j)].fill = pale_blue_fill
				front_ws["{}{}".format(letter, j)].number_format = '0.00'
		if i in range(4, 3 + 6 * num_vers, 2):
			for j in range(5, front_ws.max_row+1):
				front_ws["{}{}".format(letter, j)].fill = pale_purple_fill
				front_ws["{}{}".format(letter, j)].number_format = '0.00'
		if i in chain(range(3, 3 + 6 * num_vers, 6), range(4, 3 + 6 * num_vers, 6)):
			front_ws["{}3".format(letter)].fill = green_fill
			front_ws["{}4".format(letter)].fill = green_fill
		if i in chain(range(5, 3 + 6 * num_vers, 6), range(6, 3 + 6 * num_vers, 6)):
			front_ws["{}3".format(letter)].fill = red_fill
			front_ws["{}4".format(letter)].fill = red_fill
		if i in chain(range(7, 3 + 6 * num_vers, 6), range(8, 3 + 6 * num_vers, 6)):
			front_ws["{}3".format(letter)].fill = blue_fill
			front_ws["{}4".format(letter)].fill = blue_fill
		if i == 1 or i in range(3, 3 + 6 * num_vers):
			front_ws.column_dimensions[letter].width = 10
		elif i == 2:
			front_ws.column_dimensions[letter].width = 40
		else:
			front_ws.column_dimensions[letter].width = 70
			for j in range(5, front_ws.max_row+1):
				front_ws["{}{}".format(letter,j)].font = small_font
				front_ws["{}{}".format(letter,j)].alignment = Alignment(wrapText=True)
		for j in range(1, front_ws.max_row+1):
			front_ws["{}{}".format(letter,j)].border = black_border

	for j in range(5, front_ws.max_row+1):
		front_ws.row_dimensions[j].height = 30
	front_ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

def add_chart_to_FP(ws, chart, x_axis, y_axis, title, order):
	# Draw a chart in the Front Page
	chart.set_categories(x_axis)
	chart.y_axis.title = y_axis
	chart.title = title
	chart.width = 3 * (ws.max_row - 5) + 5
	chart.height = 16
	ws.add_chart(chart, "B{}".format(ws.max_row + 3 + 30 * order))

def plot_graph(front_ws):
	# Draw "number of apr versions" * 2 graphs on the front page
	mean_chart = BarChart3D()
	stdev_chart = BarChart3D()

	num_vers = len(apr_versions)
	for i in range(num_vers):
		VMS_start_column = 3 + i * 6
		values = Reference(front_ws, min_col=VMS_start_column, min_row=5, max_col=VMS_start_column, max_row=front_ws.max_row)
		series = Series(values, title="VMS: " + apr_versions[i])
		mean_chart.append(series)
		values = Reference(front_ws, min_col=VMS_start_column + 1, min_row=5, max_col=VMS_start_column + 1, max_row=front_ws.max_row)
		series = Series(values, title="VMS: " + apr_versions[i])
		stdev_chart.append(series)
	for i in range(num_vers):
		IPC_start_column = 5 + i * 6
		values = Reference(front_ws, min_col=IPC_start_column, min_row=5, max_col=IPC_start_column, max_row=front_ws.max_row)
		series = Series(values, title="IPC: " + apr_versions[i])
		mean_chart.append(series)
		values = Reference(front_ws, min_col=IPC_start_column + 1, min_row=5, max_col=IPC_start_column + 1, max_row=front_ws.max_row)
		series = Series(values, title="IPC: " + apr_versions[i])
		stdev_chart.append(series)
	for i in range(num_vers):
		MKB_start_column = 7 + i * 6
		values = Reference(front_ws, min_col=MKB_start_column, min_row=5, max_col=MKB_start_column, max_row=front_ws.max_row)
		series = Series(values, title="IPC: " + apr_versions[i])
		mean_chart.append(series)
		values = Reference(front_ws, min_col=MKB_start_column + 1, min_row=5, max_col=MKB_start_column + 1, max_row=front_ws.max_row)
		series = Series(values, title="IPC: " + apr_versions[i])
		stdev_chart.append(series)

	x_axis = Reference(front_ws, min_col=2, min_row=5, max_col=2, max_row=front_ws.max_row)
	y_axis = 'Time(Secs)'
	
	title = " :: " + " vs ".join(apr_versions) + " - Uptime 1 hour"

	add_chart_to_FP(front_ws, mean_chart, x_axis, y_axis, "Mean" + title, 0)
	add_chart_to_FP(front_ws, stdev_chart, x_axis, y_axis, "STDEV" + title, 1)

def write_front_page(front_ws, data_ws, maximums):
	write_header_FP(front_ws)
	write_front_page_data(front_ws, data_ws, maximums)
	format_FP(front_ws)
	plot_graph(front_ws)
############################### front page functions ################################

################################### main function ###################################
def write_report():
	# extract records from log files
	print("Data parsing in progress..."); sys.stdout.flush()

	IPC_records, VMS_records, MKB_records = collect_from_csv("logs")

	# trim each test case to cutoff_length records
	VMS_records = trim_data(VMS_records, cutoff_length)
	IPC_records = trim_data(IPC_records, cutoff_length)
	MKB_records = trim_data(MKB_records, cutoff_length)

	# Check data and pad if necessary
	VMS_records, IPC_records, MKB_records, maximums = check_and_pad(VMS_records, IPC_records, MKB_records)

	# testing purpose only
	# see_counters(VMS_records, IPC_records, MKB_records, maximums)

	# sort each list in each dict by the first column
	for apr_version in apr_versions:
		IPC_records[apr_version] = sorted(IPC_records[apr_version], key=lambda elem: elem[0])
		VMS_records[apr_version] = sorted(VMS_records[apr_version], key=lambda elem: elem[0])
		MKB_records[apr_version] = sorted(MKB_records[apr_version], key=lambda elem: elem[0])

	print("Report generation in progress..."); sys.stdout.flush()
	# write to the report file
	wb = Workbook()
	front_ws = wb.active
	front_ws.title = "Dashboard"
	data_ws = wb.create_sheet(title = 'PerformanceDATA')

	write_report_data_page(data_ws, IPC_records, VMS_records, MKB_records)
	write_front_page(front_ws, data_ws, maximums)
	
	if not os.path.exists("Report Output"):
		os.makedirs("Report Output")

	wb.save('Report Output/output_report.xlsx')
	print("Finished!"); sys.stdout.flush()

################################### main function ###################################
if __name__ == '__main__':
    write_report()